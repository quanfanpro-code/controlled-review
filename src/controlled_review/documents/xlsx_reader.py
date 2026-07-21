"""XLSX 双视图读取器。

通过两次加载工作簿（保留公式视图与缓存值视图），
提取工作表结构、单元格公式与值、隐藏行、合并区域、命名区域、打印区域，
并检测宏、数据透视表、外部连接等风险，形成 RiskNode（不执行它们）。
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from .models import (
    CellNode,
    ColumnNode,
    RiskNode,
    RowNode,
    SheetNode,
    WorkbookNode,
)


def _is_formula(value) -> bool:
    """判断值是否为公式字符串（以 = 开头）。"""
    return isinstance(value, str) and value.startswith("=")


def _build_cell(address: str, formula_cell, value_cell) -> CellNode:
    """由两个视图的单元格构造 CellNode。

    formula_cell: data_only=False 视图下的单元格（含公式）
    value_cell:   data_only=True 视图下的单元格（含缓存值）
    """
    raw_value = formula_cell.value
    formula = raw_value if _is_formula(raw_value) else None
    return CellNode(
        address=address,
        formula=formula,
        value=value_cell.value,
        raw_value=raw_value,
        number_format=formula_cell.number_format,
        data_type=formula_cell.data_type,
    )


def _build_sheet(ws_formula, ws_value, risks: list[RiskNode]) -> SheetNode:
    """由两个视图的工作表构造 SheetNode，并追加检测到的风险。"""
    cells: dict[str, CellNode] = {}
    # 遍历公式视图中的所有单元格，保持地址顺序
    for row in ws_formula.iter_rows():
        for cell in row:
            address = cell.coordinate
            value_cell = ws_value[address]
            cells[address] = _build_cell(address, cell, value_cell)

    # 行节点（仅记录显式声明过维度信息的行，含 hidden 状态）
    rows: dict[int, RowNode] = {}
    for row_idx, dim in ws_formula.row_dimensions.items():
        rows[row_idx] = RowNode(index=row_idx, hidden=bool(dim.hidden))

    # 列节点（仅记录显式声明过维度信息的列，含 hidden 状态）
    columns: dict[str, ColumnNode] = {}
    for col_letter, dim in ws_formula.column_dimensions.items():
        columns[col_letter] = ColumnNode(index=col_letter, hidden=bool(dim.hidden))

    # 合并区域
    merged_ranges = [str(r) for r in ws_formula.merged_cells.ranges]

    # 打印区域
    print_area = ws_formula.print_area

    # 工作表级命名区域
    sheet_named_ranges: list[str] = []
    if hasattr(ws_formula, "defined_names"):
        try:
            sheet_named_ranges = list(ws_formula.defined_names)
        except Exception:
            sheet_named_ranges = []

    # 数据透视表检测（openpyxl 内部属性 _pivots）
    pivots = getattr(ws_formula, "_pivots", None) or []
    if pivots:
        risks.append(
            RiskNode(
                kind="pivot_table",
                description=f"工作表「{ws_formula.title}」包含 {len(pivots)} 个数据透视表",
            )
        )

    state = ws_formula.sheet_state or "visible"
    return SheetNode(
        name=ws_formula.title,
        cells=cells,
        rows=rows,
        columns=columns,
        visible=(state == "visible"),
        state=state,
        merged_ranges=merged_ranges,
        print_area=print_area,
        named_ranges=sheet_named_ranges,
    )


def _extract_external_links(wb: Workbook) -> list[str]:
    """从工作簿提取外部链接信息（不执行，仅记录标识）。"""
    links = getattr(wb, "_external_links", None) or []
    result: list[str] = []
    for idx, link in enumerate(links, start=1):
        # link 通常是 ExternalLink 对象；尝试获取其内部引用信息
        name = getattr(link, "external_link", None) or getattr(link, "id", None) or f"external_link_{idx}"
        result.append(str(name))
    return result


def build_workbook_node(formulas: Workbook, values: Workbook, path: str = "") -> WorkbookNode:
    """由公式视图与缓存值视图构造 WorkbookNode。

    保存工作表顺序与可见状态、单元格地址/公式/缓存值/数字格式/数据类型、
    隐藏行、隐藏列、合并区域、命名区域、打印区域、外部链接；
    检测宏、数据透视表、外部连接，形成 RiskNode（不执行）。
    """
    risks: list[RiskNode] = []

    # 宏检测：openpyxl 加载 .xlsm 时若包含 VBA 工程，vba_archive 非 None
    if getattr(formulas, "vba_archive", None) is not None:
        risks.append(
            RiskNode(kind="macro", description="工作簿包含 VBA 宏工程")
        )

    # 外部链接提取（不执行）
    external_links = _extract_external_links(formulas)
    if external_links:
        risks.append(
            RiskNode(
                kind="external_link",
                description=f"工作簿包含 {len(external_links)} 个外部链接",
            )
        )

    # 工作簿级命名区域
    named_ranges: list[str] = []
    try:
        named_ranges = list(formulas.defined_names)
    except Exception:
        named_ranges = []

    # 遍历所有工作表（保持原顺序）
    sheets: list[SheetNode] = []
    for ws_formula in formulas.worksheets:
        ws_value = values[ws_formula.title]
        sheets.append(_build_sheet(ws_formula, ws_value, risks))

    return WorkbookNode(
        path=path,
        sheets=sheets,
        risks=risks,
        named_ranges=named_ranges,
        external_links=external_links,
    )


class XlsxReader:
    """XLSX 读取器。

    通过双视图加载（data_only=False 保留公式，data_only=True 保留缓存值），
    组装出不可变的 WorkbookNode 供复核流程使用。
    """

    def read(self, path: Path) -> WorkbookNode:
        formulas = load_workbook(path, data_only=False, read_only=False, keep_links=True)
        values = load_workbook(path, data_only=True, read_only=False, keep_links=True)
        return build_workbook_node(formulas, values, str(path))
