"""XLSX 测试工厂。

提供 xlsx_fixture pytest fixture，用于在临时目录中创建测试用的 XLSX 文件。
支持指定工作表名称、单元格公式/值、隐藏行等。
"""

from pathlib import Path

import pytest
from openpyxl import Workbook


@pytest.fixture
def xlsx_fixture(tmp_path):
    """返回一个工厂函数，用于创建 XLSX 文件。

    工厂参数：
        sheets: dict[str, dict[str, Any]] - 工作表名 -> {地址: 值/公式}
        hidden_rows: dict[str, list[int]] - 工作表名 -> 隐藏行号列表

    返回：XLSX 文件路径（Path）。
    """

    def _create(sheets: dict, hidden_rows: dict | None = None) -> Path:
        wb = Workbook()
        # 删除默认工作表
        wb.remove(wb.active)
        for sheet_name, cells in sheets.items():
            ws = wb.create_sheet(title=sheet_name)
            for address, value in cells.items():
                ws[address] = value
            # 设置隐藏行
            if hidden_rows and sheet_name in hidden_rows:
                for row_idx in hidden_rows[sheet_name]:
                    ws.row_dimensions[row_idx].hidden = True
        path = tmp_path / "test.xlsx"
        wb.save(path)
        return path

    return _create
