"""单元测试共享 fixtures。

将 tests/fixtures 中的 fixture 注册到 pytest，使单元测试可使用。
"""

import pytest

from tests.fixtures.docx_factory import docx_fixture, docx_with_image_table
from tests.fixtures.xlsx_factory import xlsx_fixture

__all__ = [
    "xlsx_fixture",
    "sample_xlsx",
    "recalculator",
    "docx_fixture",
    "docx_with_image_table",
]


@pytest.fixture
def sample_xlsx(tmp_path):
    """创建一个简单的 XLSX 文件用于测试。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = "=SUM(A1:A2)"
    path = tmp_path / "sample.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def workbook_nodes():
    """创建测试用工作簿节点列表。

    提供三张工作表，每张都填充可被多信号分类识别的单元格：
    - 标题单元格（A1）保存工作表名称
    - 单位标注（A2）"单位：千元" -> CNY_THOUSAND
    - 期间列头（B3="本期"，C3="上期"）-> (current, prior)
    - 行项目（A4 起）提供该报表的典型项目
    - "合并资产负债表1"（合并层面，拆分第一部分）
    - "合并资产负债表2"（合并层面，拆分第二部分）
    - "母公司资产负债表"（母公司层面）
    """
    from controlled_review.documents.models import CellNode, SheetNode, WorkbookNode

    def make_sheet(name: str, line_items: tuple[str, ...] | None = None) -> SheetNode:
        """构造单张工作表节点，预置标题/单位/期间/行项目单元格。"""
        cells: dict[str, CellNode] = {}
        # 标题单元格
        cells["A1"] = CellNode(
            address="A1",
            formula=None,
            value=name,
            raw_value=name,
            number_format=None,
            data_type="s",
        )
        # 单位标注
        cells["A2"] = CellNode(
            address="A2",
            formula=None,
            value="单位：千元",
            raw_value="单位：千元",
            number_format=None,
            data_type="s",
        )
        # 期间列头
        cells["B3"] = CellNode(
            address="B3",
            formula=None,
            value="本期",
            raw_value="本期",
            number_format=None,
            data_type="s",
        )
        cells["C3"] = CellNode(
            address="C3",
            formula=None,
            value="上期",
            raw_value="上期",
            number_format=None,
            data_type="s",
        )
        # 行项目
        if line_items:
            for i, item in enumerate(line_items, start=4):
                addr = f"A{i}"
                cells[addr] = CellNode(
                    address=addr,
                    formula=None,
                    value=item,
                    raw_value=item,
                    number_format=None,
                    data_type="s",
                )
        return SheetNode(name=name, cells=cells)

    sheets = [
        make_sheet("合并资产负债表1", line_items=("货币资金", "应收账款")),
        make_sheet("合并资产负债表2", line_items=("存货", "固定资产")),
        make_sheet("母公司资产负债表", line_items=("货币资金", "实收资本")),
    ]
    return [WorkbookNode(path="test.xlsx", sheets=sheets)]


@pytest.fixture
def recalculator(monkeypatch):
    """返回 mock 了 Excel 的 OfficeRecalculator。

    用 unittest.mock.MagicMock 模拟 Excel COM 对象，
    单元测试不依赖真实 Office 安装。

    mock 行为对齐真实 Excel：
    - mock_excel.ActiveWorkbook 返回 mock_workbook（与真实 Open 失败回退路径一致）
    - mock_excel.Workbooks.Open 返回 mock_workbook（保留旧约定）
    mock_excel 与 mock_workbook 附加为实例属性 _mock_excel / _mock_workbook，
    供测试验证关键调用（_FlagAsMethod、AutomationSecurity、Close 等）。
    """
    from unittest.mock import MagicMock

    from controlled_review.documents.office_recalc import OfficeRecalculator

    mock_excel = MagicMock()
    mock_workbook = MagicMock()
    # 对齐真实 Excel 行为：ActiveWorkbook 返回工作簿对象
    mock_excel.ActiveWorkbook = mock_workbook
    mock_excel.Workbooks.Open.return_value = mock_workbook
    monkeypatch.setattr(
        "controlled_review.documents.office_recalc.dispatch_excel",
        lambda: mock_excel,
    )
    recalculator = OfficeRecalculator()
    # 暴露 mock 给测试，验证关键调用
    recalculator._mock_excel = mock_excel
    recalculator._mock_workbook = mock_workbook
    return recalculator
